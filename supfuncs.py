import datetime
import psycopg2
import os
import pandas as pd
import numpy as np
from jinjasql import JinjaSql
from sqlalchemy import create_engine

from openpyxl import Workbook
from openpyxl.styles import colors, Font, Color, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

nan = np.nan

# Formatting Variables

# Create Alignment
center = Alignment(horizontal = 'center')
left = Alignment(horizontal = 'left')
right = Alignment(horizontal = 'right')

# Create Font
titleFont = Font(size=20, bold=True)
boldFont = Font(bold=True)
headingFont = Font(size=16, bold=True)

# *************************** #
# Report Functions
# *************************** #

def build_project_report(project_id):
    """
    This will return the completed project report as a wb, 
    ready to send to user.
    """

    projDF, tsDF, roomsDF = get_proj_details(project_id)
    
    idx = 0
    wb = create_details_page(idx, projDF)
    
    wb = add_labour_to_workbook(idx, wb, projDF, tsDF)
    
    wb = add_rooms_to_workbook(wb, roomsDF)

    return wb


def get_proj_details(project_id):
    """
    return projectDF and timesheetDF for the specified project_id

    For on demand, project specific report building
    """
    # Create a connection
    postgres_connection_URL = os.environ.get('DATABASE_URL')
    engine = create_engine(postgres_connection_URL)

    # Open connection with a context manager
    # run a query and save it in a variable
    with engine.connect() as cnx:
        project_query = "SELECT * FROM Project WHERE id=" + str(project_id)
        projects = pd.read_sql_query(project_query, cnx)

        sites = pd.read_sql_query("SELECT id, name, code FROM Site", cnx)
        areas = pd.read_sql_query("SELECT * FROM Area", cnx)
        status = pd.read_sql_query("SELECT * FROM Status", cnx)
        typeOfWork = pd.read_sql_query("SELECT * FROM Worktype", cnx)
        users = pd.read_sql_query('SELECT id, username FROM "user"', cnx)
        
        timesheet_query = "SELECT * FROM timesheet where project_id=" + str(project_id)
        timesheets = pd.read_sql_query(timesheet_query, cnx)
        
        clients = pd.read_sql_query("SELECT id, name FROM client", cnx)
    
        room_query = "SELECT bm_id, name, location, description, date_last_paint, freq, site_id, area_id, orig_paint_date FROM room WHERE room.id IN(SELECT room_id FROM project_room WHERE project_id=" + str(project_id) + ");"
        roomsDF = pd.read_sql_query(room_query, cnx)

    # Change to project_id
    projectDF = projects.rename({'id':'project_id'},axis=1)
    
    # Rename site columns for merge
    siteDF = sites.rename({'name':'site_name', 'code':'site_code'},axis=1)
    
    # Merge sites to projects
    projectDF = projectDF.merge(siteDF, left_on="site_id", right_on="id")
    # drop the now un necessary columns
    projectDF = projectDF.drop(['id', 'site_id'], axis=1)
    
    # Rename client columns for merge
    clientDF = clients.rename({'name':'client_name'},axis=1)
    
    # Merge clientDF on Projects
    projectDF = projectDF.merge(clientDF, left_on="client_id", right_on="id")
    projectDF = projectDF.drop(["id","client_id"], axis=1)
    
    statusDF = status.rename({"name":"status"}, axis=1)
    statusDF = statusDF[['id', 'status']]
    
    # Merge Status on Project
    projectDF = projectDF.merge(statusDF, left_on="status_id", right_on="id")
    projectDF = projectDF.drop(['id', 'status_id'], axis=1)
    
    
    typeDF = typeOfWork.rename({"name":"worktype", "code":"type_code"}, axis=1)
    typeDF = typeDF[['id','worktype','type_code']]
    
    projectDF = projectDF.merge(typeDF, left_on="typeOfWork_id", right_on="id")
    projectDF = projectDF.drop(['id', 'typeOfWork_id'], axis=1)
    
    timesheetDF = timesheets.rename({"id":"timesheet_id", "completed":"day_completed"}, axis=1)
    timesheetDF = timesheetDF.merge(users, left_on='user_id', right_on='id')
    timesheetDF = timesheetDF.drop(['id', 'user_id', 'isNotWorkDay'], axis=1)
    
    projectForTimesheetDF = projectDF[['project_id', 'name', 'site_code', 'worktype']]
    projectForTimesheetDF = projectForTimesheetDF.rename({"name":"project_name"})
    
    timesheetDF = timesheetDF.merge(projectForTimesheetDF, on="project_id")

    tsDF = timesheetDF[['dateOfWork', 'username', 'hours', 'comment', 'name', 'project_id', 'site_code', 'dateSubmit']]

    totalHrs = tsDF.hours.sum()

    projectDF['totalHours'] = totalHrs

    return projectDF, tsDF, roomsDF


def create_details_page(idx, projectDF):
    # Create a Project Details Page Template
    wb = Workbook()
    ws = wb.active
    ws.title = 'ProjectDetails'

    # Enter the label fields
    ws['A2'] = 'Client Name:'
    ws['A3'] = 'Site Name:'
    ws['A4'] = 'Work Type:'
    ws['A5'] = 'Project Status:'

    ws['A7'] = 'Purchase Order #:'
    ws['A8'] = 'Work Order #:'

    ws['A10'] = 'Description:'
    ws['A11'] = 'Damage Comment:'
    ws['A12'] = 'Exten Circumstances:'

    ws['A14'] = 'Start Date:'
    ws['A15'] = 'Finish Date:'
    ws['A16'] = 'Days to Complete:'

    ws['A18'] = 'Quoted Amount:'
    ws['A19'] = 'Invoiced Amount:'

    ws['A21'] = 'Labour'
    ws['A22'] = 'Estimated Hours:'
    ws['A23'] = 'Actual Hours:'

    ws['A25'] = 'Total Cost of Labour:'

    ws['A27'] = 'Materials' 
    ws['A29'] = 'Total Cost of Materials'
    
    ws['A35'] = 'Report Generated on:'

    # Define which cells we want to contain which data
    nameCell = ws['A1']
    matCell = ws['A27']
    labCell = ws['A21']

    clientNameCell = ws['B2']
    siteCell = ws['B3']
    workTypeCell = ws['B4']
    statusCell = ws['B5']
    POCell = ws['B7']  
    WOCell = ws['B8']  
    descCell = ws['B10'] 
    dmgCell = ws['B11'] 
    excircCell = ws['B12'] 
    stdateCell = ws['B14'] 
    fndateCell = ws['B15'] 
    daystcCell = ws['B16'] 
    quoteCell = ws['B18'] 
    invCell = ws['B19']
    esthourCell = ws['B22']
    acthourCell = ws['B23']
    labcostCell = ws['B25']
    matcostCell = ws['B29']
    
    genDateCell = ws['B35']
    
    # Set Values of those Cells
    nameCell.value = projectDF.iloc[idx]['name']
    clientNameCell.value = projectDF.iloc[idx]['client_name']
    siteCell.value = projectDF.iloc[idx]['site_name']
    workTypeCell.value = projectDF.iloc[idx]['worktype']
    statusCell.value = projectDF.iloc[idx]['status']
    POCell.value = projectDF.iloc[idx]['purchase_order']
    WOCell.value = projectDF.iloc[idx]['work_order']
    descCell.value = projectDF.iloc[idx]['description']
    dmgCell.value = projectDF.iloc[idx]['damage_comment']
    excircCell.value = projectDF.iloc[idx]['extenuating_circumstances']

    if projectDF.loc[[idx],'date_start'].isna().all():
        stdateCell.value = 'Not Started'
    else: 
        stdateCell.value = projectDF.iloc[idx]['date_start'].date() # Use this for the date as an object
    #     dateStr = projectDF.iloc[6]['date_start'].date().strftime('%B %d, %Y')
    #     stdateCell.value = dateStr

    if projectDF.loc[[idx],'date_finished'].isna().all():
        fndateCell.value = 'Still in Progress'
    else: 
        fndateCell.value = projectDF.iloc[idx]['date_finished'].date() # Use this for the date as an object
    #     dateStr = projectDF.iloc[idx]['date_finished'].date().strftime('%B %d, %Y')
    #     fndateCell.value = dateStr

    # Calculate the number of days that the project has been in progress for
    if projectDF.iloc[idx]['date_finished']:
        daysInProgress = projectDF.iloc[idx]['date_finished'] - projectDF.iloc[idx]['date_start']
    else:
        daysInProgress = datetime.datetime.now() - projectDF.iloc[idx]['date_start']
        ws['A16'] = 'Days to In Progress:'
    daystcCell.value = daysInProgress.days

    quoteCell.value = projectDF.iloc[idx]['quote_amt']
    invCell.value = projectDF.iloc[idx]['invoice_amt']

    esthourCell.value = projectDF.iloc[idx]['hours_estimate']
    acthourCell.value = projectDF.iloc[idx]['totalHours']
    # labcostCell.value = 
    # matcostCell.value = 
    
    genDateCell.value = datetime.datetime.now().strftime('%A %B %d, %Y')
        
    # Formatting
    # Merge Title Row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    # Labour
    ws.merge_cells(start_row=21, start_column=1, end_row=21, end_column=2)
    # Materials
    ws.merge_cells(start_row=27, start_column=1, end_row=27, end_column=2)

    # Apply alignment
    # I want all of col B left aligned and bold
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=2).alignment = left
        ws.cell(row=row, column=1).alignment = right
        ws.cell(row=row, column=1).font = boldFont

    nameCell.alignment = center
    labCell.alignment = center
    matCell.alignment = center

    # Apply font
    nameCell.font = titleFont
    labCell.font = headingFont
    matCell.font = headingFont

    # Apply date format
    stdateCell.number_format = 'MMM DD, YYYY'
    fndateCell.number_format = 'MMM DD, YYYY'

    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=26)

    ws.column_dimensions = dim_holder

    return wb


def add_labour_to_workbook(idx, wb, projectDF, tsDF):
    ws2 = wb.create_sheet(title="LabourDetails")
    pid = projectDF.iloc[idx]['project_id']
    pidtsDF = tsDF.loc[tsDF.project_id==pid]
    for r in dataframe_to_rows(pidtsDF, index=False, header=True):
        ws2.append(r)
    
    # Format the labour details
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 9 
    ws2.column_dimensions['C'].width = 9
    ws2.column_dimensions['D'].width = 45
    ws2.column_dimensions['E'].width = 30
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 9
    ws2.column_dimensions['H'].width = 19

    for col in range(ws2.min_column, ws2.max_column+1):
        ws2.cell(row=1, column=col).font = boldFont
    
    return wb


def add_rooms_to_workbook(wb, roomsDF):
    from openpyxl.utils.dataframe import dataframe_to_rows

    roomsDF = roomsDF.fillna('NA')

    ws3 = wb.create_sheet(title="Rooms")
    for r in dataframe_to_rows(roomsDF, index=False, header=True):
        ws3.append(r)
    
    # Format the labour details
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 30
    ws3.column_dimensions['E'].width = 18
    ws3.column_dimensions['F'].width = 9
    ws3.column_dimensions['G'].width = 9
    ws3.column_dimensions['H'].width = 9
    ws3.column_dimensions['I'].width = 18

    for col in range(ws3.min_column, ws3.max_column+1):
        ws3.cell(row=1, column=col).font = boldFont
    
    for row in range(ws3.min_row+1, ws3.max_row+1):
        ws3.cell(row=row, column=5).number_format = 'MMM DD, YYYY'
        ws3.cell(row=row, column=5).alignment = right
        ws3.cell(row=row, column=9).number_format = 'MMM DD, YYYY'
        ws3.cell(row=row, column=9).alignment = right

    return wb


def build_upcomingDF(site_id):
    """
    Create a Pandas DataFrame of the upcoming rooms within the select site_id
    """

    postgres_connection_URL = os.environ.get('DATABASE_URL')
    engine = create_engine(postgres_connection_URL)
    
    template = """SELECT *
    FROM room
    WHERE site_id = {{site_id}}
    """
    data = {
    "site_id": site_id
    }
    j = JinjaSql()
    query, bind_params = j.prepare_query(template, data)

    rooms = pd.read_sql(sql=query, con=engine, params=bind_params)

    # drop any rooms that have never been painted
    rooms = rooms.loc[~rooms.date_last_paint.isna()]

    # Convert to datetime
    rooms.date_last_paint = pd.to_datetime(rooms.date_last_paint)
    rooms.date_next_paint = pd.to_datetime(rooms.date_next_paint)

    today = datetime.date.today()
    for row in rooms.index:
        try:
            date_diff = rooms.loc[row, 'date_next_paint'].date() - today
            years_till_paint = round(date_diff.days/365, 1)
        except TypeError:
            years_till_paint = nan
        if rooms.loc[row, 'freq'] == -1:
            rooms.loc[row, 'due_in_X_years'] = nan
        else:
            rooms.loc[row, 'due_in_X_years'] = years_till_paint

    rooms = rooms.loc[rooms.due_in_X_years <=1].copy()

    rooms.sort_values('date_next_paint', inplace=True)

    rooms = rooms[['bm_id', 'name', 'date_last_paint', 'freq', 'date_next_paint', 'due_in_X_years']].copy()

    rooms.rename({'bm_id':'BM ID', 'name':'Room Name', 'date_last_paint':'Date of last Painting', 'freq':'Frequency', 'date_next_paint':'Next Painting', 'due_in_X_years':'Years till due'}, axis=1, inplace=True)

    rooms.reset_index(drop=True, inplace=True)

    return rooms


def build_asneededDF(site_id, area_id=None):
    """
    Create a Pandas DataFrame of the as needed rooms within the select site_id
    """

    postgres_connection_URL = os.environ.get('DATABASE_URL')
    engine = create_engine(postgres_connection_URL)
    
    # Potentially add an area filter as well?
    # if area_id:
    #     template with area_id
    # else
    #     template

    template = """SELECT *
    FROM room
    WHERE site_id = {{site_id}}
    AND freq = -1"""
    
    data = {
    "site_id": site_id
    }

    j = JinjaSql()
    query, bind_params = j.prepare_query(template, data)
    
    # Make DF of just as needed rooms for selected site.
    as_neededDF = pd.read_sql(sql=query, con=engine, params=bind_params)
    as_neededDF = as_neededDF[['bm_id', 'name', 'date_last_paint']].copy()

    # split dated from non-dated
    no_dateDF = as_neededDF.loc[as_neededDF.date_last_paint.isna()].copy()
    datedDF = as_neededDF.loc[~as_neededDF.date_last_paint.isna()].copy()

    # Set to datetime, and drop the time part.
    datedDF.date_last_paint = datedDF.date_last_paint.astype('datetime64[ns]')
    datedDF.date_last_paint = datedDF.date_last_paint.dt.strftime('%Y-%m-%d')

    # Sort them by date and bm_id
    no_dateDF.sort_values(['bm_id'],inplace=True)
    datedDF.sort_values(['date_last_paint','bm_id'],ascending=True, inplace=True)

    sortedDF = no_dateDF.append(datedDF)
    sortedDF.rename({'bm_id':'BM ID', 'name':'Room Name', 'date_last_paint':'Date of last Painting'}, axis=1, inplace=True)
    # sortedDF.set_index('BM ID', inplace=True)
    sortedDF.reset_index(drop=True, inplace=True)

    return sortedDF


def build_filtered_project_list(client_id, site_id, status_id, typeOfWork_id, start_date_after, start_date_before, finish_date_after, finish_date_before):
    """
    Build list of projects based on filter input
    """
    # Establish Connection
    postgres_connection_URL = os.environ.get('DATABASE_URL')
    engine = create_engine(postgres_connection_URL)
    # Get list of all projects
    all_projects=pd.read_sql_query("SELECT * FROM Project", con=engine)

    filt_dict = {
    "client_id": client_id, 
    "site_id": site_id, 
    "status_id": status_id, 
    "typeOfWork_id": typeOfWork_id, 
}
    # handle no param case
    if client_id==None and site_id==None and status_id==None and typeOfWork_id==None:
        res = all_projects
    else:
        # Build query string that only queries for the entered params
        query_string = ''
        for key, val in filt_dict.items():
            if val != None:
                if query_string != '':
                    query_string += " & "
                query_string += (str(key) + " == " + str(val))
        # filter results 
        res = all_projects.query(query_string)

    print(client_id)
    print(site_id) 
    print(status_id)
    print(typeOfWork_id)
    print(start_date_after)
    print(start_date_before)
    print(finish_date_after)
    print(finish_date_before)

    # filter on start date
    # 4 possible combos
    if start_date_after and start_date_before==None:
        start_date_after = np.datetime64(start_date_after)
        start_date_mask = (res['date_start'] > start_date_after)
        res = res.loc[start_date_mask]
    elif start_date_before and start_date_after==None:
        start_date_before = np.datetime64(start_date_before)
        start_date_mask = (res['date_start'] <= start_date_before)
        res = res.loc[start_date_mask]
    elif start_date_after and start_date_before:
        start_date_before = np.datetime64(start_date_before)
        start_date_after = np.datetime64(start_date_after)
        start_date_mask = (res['date_start'] > start_date_after) & (res['date_start'] <= start_date_before)
        res = res.loc[start_date_mask]
    else: # both are None
        pass
    
    # filter on end date
    # 4 possible combos
    if finish_date_after and finish_date_before==None:
        finish_date_after = np.datetime64(finish_date_after)
        finish_date_mask = (res['date_finished'] > finish_date_after)
        res = res.loc[finish_date_mask]
    elif finish_date_before and finish_date_after==None:
        finish_date_before = np.datetime64(finish_date_before)
        finish_date_mask = (res['date_finished'] <= finish_date_before)
        res = res.loc[finish_date_mask]
    elif finish_date_after and finish_date_before:
        finish_date_after = np.datetime64(finish_date_after)
        finish_date_before = np.datetime64(finish_date_before)
        finish_date_mask = (res['date_finished'] > finish_date_after) & (res['date_finished'] <= finish_date_before)
        res = res.loc[finish_date_mask]
    else: # both are None
        pass

    # Output the final list of Ids
    project_id_list = res.id.to_list()
    return project_id_list